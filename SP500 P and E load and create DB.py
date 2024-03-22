# SP500 index and earnings analysis
# Projections of earnings and operating margins
# initiated:  2023 08
# current v:  2023 09 01

import os
import pickle
import sys
from datetime import datetime
from pprint import pprint

import openpyxl.utils.cell
import pandas as pd
from openpyxl import load_workbook

#######################  Parameters  ##################################

"""set parameters for the program

   creates or updates earnings, p/e, and margin data from
   'sp-500-eps-est ... .xlsx' files in 'sp500_pe_dict'.
   pickles 'sp500_pe_dict' in 'sp500_pe_data.pickle'

file structure
|
|__ sp_data_dir/
|   |__ <sp-500 .xlsx data docs>
|
|__ Python_code_dir/
    |__ <this program>
    |__ output_dir/
|       |__<sp500_pe_data.pickle>

"""

data_dir = "/Users/.../sp_data_dir"
code_dir = "/Users/.../Python_code_dir"

 # set working directory to the code directory
os.chdir(code_dir)
output_dir = f"{code_dir}/output_dir"
output_file = 'sp500_pe_data.pickle'

column_names = ['date', 'price', 'op_eps', 'rep_eps',
                'op_p/e', 'rep_p/e', '12m_op_eps', '12m_rep_eps']
columns_mdy = ['month', 'day', 'year']
columns_qtrly = ['date', 'div_ps', 'sales_ps',
                 'bk_val_ps', 'capex_ps', 'price', 'divisor']

wksht_name = 'ESTIMATES&PEs'
actual_key = 'ACTUALS'
actual_key2 = 'Actuals'
proj_key = 'ESTIMATES'
margin_key = 'QTR'

rr_file_addr = data_dir + "/DFII10.xlsx"
rr_col_name = 'real_int_rate'

qtr_to_date_dict = {'Q1': '03-31',
                    'Q2': '06-30',
                    'Q3': '09-30',
                    'Q4': '12-31'}

date_to_qtr_dict = dict(zip(qtr_to_date_dict.values(),
                            qtr_to_date_dict.keys()))

qtrly_wksht_name = 'QUARTERLY DATA'
qtrly_key = 'END'
last_col_qtrly = 'I'

empty_cols = [4, 7]
empty_cols_qtr = [2, 3]
max_to_read = 1000
first_col = 'A'
last_col ='J'

'''
sp500_pe_dict --
        {'history': 
                    {'date': date as string
                     'src_file' : name of data file
                     'actuals': dataframe, with op_margins, etc,
                                10-yr tips
                    }
        'projections':
                    {name_date: key is date of projection as str
                                {'estimates': dataframe
                                 'price': float actual price on the date
                                }
                    }
        'input_xlsx_set': { } #set of files already read from data_dir
        }
'''

# ? read .toml file here
#######################################################################

#######################  MAIN Function  ###############################

def main():
    '''create or update earnings, p/e, and margin data
       from 'sp-500-eps-est ...' files in 'sp500_pe_dict'
       
       pickle 'sp500_pe_dict' in data_dir/output_dir
    
    '''

# +++++  read the data from the files in the data directory  ++++++++++
# read 'sp500_pe_data.pickle', if qny, and save sp500_pe-dict as backup
    if output_file in os.listdir(output_dir):    # found previous output file
        with open(f"{output_dir}/{output_file}", "rb") as file:
            sp500_pe_dict = pickle.load(file)

        # files previously read & current source file for history
        input_xlsx_set = sp500_pe_dict.get('input_xlsx_set')
        src_file = sp500_pe_dict['history'].get('src_file')
        
        print('\n====================================')
        print('Latest file previously read')
        print(max(sp500_pe_dict['input_xlsx_set']))
        print(f'History source file is {src_file}')
        print('====================================')
        
        # used below to update real rates
        prev_update = float(src_file[15:19]) * 10000 + \
                      float(src_file[20:22]) * 100 + \
                      float(src_file[23:25])
        
    else:       # if no previous output file
        # initialize set of previously read xlsx to be empty
        print('\n====================================')
        print(f"No history file in {output_dir}")
        print(f"Creating new history file in {data_dir}")
        print('====================================')
        
         # used below to update real rates
        prev_update = 0.0
        
        # initialize dict
        sp500_pe_dict = dict.fromkeys(('history', 'projections', 
                                       'input_xlsx_set'))
        input_xlsx_set = set()
        sp500_pe_dict['input_xlsx_set'] = input_xlsx_set
        src_file = None

# +++++  fetch the list of data files to be read  +++++++++++++++++++++
# https://www.geeksforgeeks.org/python-list-files-in-a-directory/
    data_files = [file for file in os.listdir(data_dir)
                  if (file.startswith('sp-500-eps') and
                     (file.endswith('.xlsx')) and
                     (file not in input_xlsx_set))]
    
    # stop processing if data_files is empty
    if not data_files:
        print('\n====================================')
        print(f'No new files in {data_dir}')
        print('All files have been read previously')
        print('====================================')
        sys.exit()
    
# +++++  sort files, decreasing dates  ++++++++++++++++++++++++++++++++
# file with the most recent date has the most complete history
# historical data from the 1st file will be read into the history df
    data_files.sort(reverse = True)
    print('\n====================================') 
    print(f'Found unread files in {data_dir}:')
    print(data_files)
    print('====================================')
    
# +++++  save the previous output file  +++++++++++++++++++++++++++++++    
    with open(f"{output_dir}/sp500_pe_dict_backup.pickle", "wb") as file:
        pickle.dump(sp500_pe_dict, file)
        
# +++++ read historical data from latest file in data_file  +++++++++++
# if this file contains new historical data, update sp500_eps_dict['history']
# otherwise, only update sp500_eps_dict['projections']
    if not (src_file) or (data_files[0] > src_file):   # new historical data

        print('\n====================================')
        print(f'Updating historical data from: {data_files[0]}')
        print('====================================')
        
        # address of the most recent file
        file_addr = f'{data_dir}/{data_files[0]}'
        
        actual_df, name_date = data_reader([actual_key], [None],
                                           file_addr, 
                                           wksht_name, first_col, 
                                           last_col, 
                                           empty_cols=empty_cols,
                                           date_key='A', 
                                           column_names=column_names,
                                           columns_mdy=columns_mdy)
        actual_df.price = actual_df.price.round(0)
        
        # include qtr and year_qtr cols in the history data
        mdy_to_qy(actual_df)
        
        # if any date is None, abort
        if (name_date is not None and
            all([item is not None
                for item in actual_df['date']])):
    
        # update sp500_pe_dict['history']['date'] and ['actuals']
            sp500_pe_dict['history'] = {'date': name_date,
                                        'actuals': actual_df}
            del actual_df
            
        else:
            print('In main(), history:')
            print(f'Abort: sp-500 {name_date} missing history date')
            sys.exit()
            
    # margins
        margins_df = margin_reader([margin_key], file_addr, 
                                   wksht_name, first_col)
        
        marg_cols_name = list(margins_df.columns)
        # id_vars contains the list of col names not to be melted
        # value vars contains the list of col names to be melted into 1 col
        margins_df = pd.melt(margins_df, id_vars=['qtr'], 
                             value_vars=marg_cols_name[1:])
        
        margins_df['date'] = \
            [f"{qtr_to_date_dict[margins_df['qtr'][idx]]}-"+ \
             f"{margins_df['variable'][idx]}"
             for idx in range(len(margins_df.index))]
            
        # tidy
        margins_df.rename(columns = {'value':'op_margin'},
                          inplace = True)
        
        # merge margins with sp500_pe_dict['history']['actuals']
        sp500_pe_dict['history']['actuals'] = pd.merge(
            sp500_pe_dict['history']['actuals'], 
            margins_df[['date', 'op_margin']], 
            how="left", on=["date"])
        del margins_df
        
    # real interest rates, from FRED DFII10
        real_rt_df = real_rt_reader(rr_file_addr,
                                    rr_col_name,
                                    prev_update)
        
        sp500_pe_dict['history']['actuals'] = pd.merge(
            sp500_pe_dict['history']['actuals'], 
            real_rt_df, 
            how= "left", on= ['year_qtr'])
        del(real_rt_df)
            
    # qtrly_data
        qtrly_df, _ = data_reader([qtrly_key], [None],
                                  file_addr, 
                                  qtrly_wksht_name, 
                                  first_col, last_col_qtrly,
                                  empty_cols=empty_cols_qtr, 
                                  column_names=columns_qtrly,
                                  columns_mdy=columns_mdy)
        
        # merge qtrly with sp500_pe_dict['history']['actuals']
        # NB double [[]] to select cols in qtrly_df
        sp500_pe_dict['history']['actuals'] = pd.merge(
            sp500_pe_dict['history']['actuals'], 
            qtrly_df[['date', 'div_ps', 'sales_ps',
                    'bk_val_ps', 'capex_ps', 'divisor']], 
            how="left", on=["date"])
        del qtrly_df
        
        sp500_pe_dict['history']['src_file'] = data_files[0]
        
    else:   # no new historical data
        print(f'No new file updates the previous historical record:')
        print(f'{src_file} is more recent than {data_files[0]}')
        print('\n')

# +++++ update projection data ++++++++++++++++++++++++++++++++++++++++
# loop through all files, fetch projections of earnings for each date
    for file in data_files:
        proj_dict = dict()
        print(file)
        file_addr = f'{data_dir}/{file}'
        
# projections of earnings
        proj_df, name_date = data_reader([proj_key], 
                                         [None, actual_key, 
                                          actual_key2],
                                         file_addr, 
                                         wksht_name, first_col, 
                                         last_col, 
                                         empty_cols=empty_cols,
                                         date_key='A',
                                         column_names=column_names,
                                         columns_mdy=columns_mdy)
        proj_df = proj_df.drop(columns=['price'])
        
        # include qtr and year_qtr cols in the history data
        mdy_to_qy(proj_df)
        
        # if date is None, abort, and continue to the next file
        if ((name_date is None) or
            (any([item is None
                  for item in proj_df['date']]))):
            print('In main(), projections:')
            print(f"Skipped sp-500 {name_date}: missing projtn date")
            continue
        
        price = round(proj_df['rep_p/e'].values[0] * 
                    proj_df['12m_rep_eps'].values[0], 0)
        
        # create key for sub dict that contains this DF
        dt_name = datetime.strptime(name_date, 
                                    "%m-%d-%Y").date()
        
        proj_dict[dt_name] = {'estimates': proj_df, 
                                'price': price}
        
        if sp500_pe_dict['projections'] is None:
            sp500_pe_dict['projections'] = proj_dict
        else:
            sp500_pe_dict['projections'][dt_name] = \
                proj_dict[dt_name]
        
        sp500_pe_dict['input_xlsx_set'].update([file])        
        print(file, '\n')
    
# +++++  pickle the new sp500_pe_dict +++++++++++++++++++++++++++++++++
    '''
    pprint(sp500_pe_dict['history']['date'])
    pprint(sp500_pe_dict['history']['src_file'])
    print(sp500_pe_dict['history']['actuals'])
    print(sp500_pe_dict['history']['actuals'].info())
    print(sp500_pe_dict['history']['actuals']['year_qtr'])
    pprint(sp500_pe_dict['projections'][name_date]['estimates'])
    print(sp500_pe_dict['projections'][name_date]['estimates'].info())
    print(sp500_pe_dict['projections'][name_date]['estimates']['year_qtr'])
    pprint(sp500_pe_dict['projections'][name_date]['price'])
    pprint(sp500_pe_dict['input_xlsx_set'])
    sys.exit()
    '''
    
    
# +++++ store the updated sp500_pe_dict
# https://realpython.com/python-pickle-module/
# https://www.geeksforgeeks.org/pickle-python-object-serialization/
    
    # save new sp500_pe-dict
    with open(f"{output_dir}/{output_file}", "wb") as file:
        pickle.dump(sp500_pe_dict, file)
    
    print('\n==============================')
    print('Retrieval is complete')
    print('==============================')
        
    #with open(f"{output_dir}/{output_file}", "rb") as file:
    #    m_dict = pickle.load( file)
    #pprint(m_dict)


#######################  Helper Functions  ############################


def read_wksht(file_addr, wksht_name, date_key=None):
    """_summary_
    This helper function returns the worksheet and date
    for the worksheet name in the workbook 
    specified by the file address
    
    Args:
        file_addr (_type_): _description_
        wksht_name (_type_): _description_
        date_key (_type_): _description_

    Returns:
        _type_: _description_
    """

    wkbk = load_workbook(filename=file_addr,
                         read_only=True,
                         data_only=True)
    wksht = wkbk[wksht_name]
    max_to_read = wksht.max_row
    
    # find cell with the date of the wkbk, fetch its value
    # crawl down col A; find first datetime entry
    if date_key is not None:
        d_row = 1
        date = wksht['A1'].value
        while (not isinstance(date, datetime) and
               (d_row < max_to_read)):
            d_row += 1
            date = wksht[f'{first_col}{d_row}'].value
        if d_row == max_to_read:
            return [wksht, None]
        name_date = date_as_string(date)
    else:
        name_date = None
                                                                          
    return [wksht, name_date]


def find_key_loc(start, stop, wksht, key_wrds, col=None, row=None):
    """
    Set key and col to search a col down its rows for a string 
    that matches key.
    Set row, not col, to search a row along its cols for None

    Args:
        start (_type_): _description_
        stop (_type_): _description_
        key (_type_): _description_
    """
    # helper function for returning step
    def return_step(item, key_wrds):
        if any(k is None for k in key_wrds):
            if ((item is None) or 
                (item in key_wrds)):
                return True
        elif ((item is not None) and 
              (item in key_wrds)):
            return True
        else:
            return False
        
    if (((col is not None) and (row is not None)) or
       ((col is None) and (row is None))):
        print('In find_key_loc():')
        print('One and only one of row, col must be None')
        print(start, stop, wksht, key_wrds, col, row)
        sys.exit()
    
    # scan rows
    if col is not None:
        for step in range(start, stop):
            item = wksht[f'{col}{step}'].value
            if isinstance(item, str):
                item = "".join(list(item.split()))
            if return_step(item, key_wrds):
                return step
    # scan cols
    elif row is not None:
        for step in range(start, stop):
            col = openpyxl.utils.cell.get_column_letter(step)
            item = wksht[f'{col}{row}'].value
            if return_step(item, key_wrds):
                return step
            
    print('In find_key_loc():')                
    print(f'key {key_wrds} not found')
    print(start, stop, wksht, key_wrds, col, row)
    sys.exit()


def data_from_sheet(wksht, start_col, stop_col, skip_cols,
                    start, stop):
    """_summary_
    
    This helper function fetches the block of data specified
    by first_col, start (row) and last_col, stop (row)
    in the specified worksheet

    Args:
        first_col (_type_): _description_
        last_col (_type_): _description_
        start (_type_): _description_
        stop (_type_): _description_
    """
    # rng references the block of data
    # the comprehense fetches the data over cols for each row
    rng = wksht[f'{start_col}{start}:{stop_col}{stop}']
    data = [[col_cell.value 
            for ind, col_cell in enumerate(row)
            if ind not in skip_cols]
            for row in rng]
    
    return data

    
def date_as_string(date):
    """
    ensure all dates are strings, 'm-d-y'
    ensure that all y contain 4 characters

    Args:
        date : datetime or string

    Returns:
        date : date
    """
    if isinstance(date, datetime):
        return f'{date.month:02}-{date.day:02}-{date.year:04}'
    # otherwise date is str
    # remove any additional expression from str date
    # extraneous expressions in xlsx follow the date, separated by a blank
    else:
        date = str(date)
        date_lst = date.split(" ")
        for item in date_lst:
            if (len(item)>7 and len(item)<11 and
                all([ch in '1234567890/' for ch in item])):
                m, d, y = item.split("/")
                if len(y) == 2:
                    if int(y) < 20:
                        return f'{m:>02}-{d:>02}-20{y:^2}'
                    else:
                        return f'{m:>02}-{d:>02}-19{y:^2}'
                else:
                    return f'{m:>02}-{d:>02}-{y:^4}'
            else:
                return None
            
            
def mdy_to_qy(df):
    ''' input date is m-d-y string for last day of qtr
        convert date string to qtr and year_qtr strings
        add these cols to df
    '''

    df['qtr'] = [f"{date_to_qtr_dict[item[0:5]]}"
                 for item in df['date']]
    df['year_qtr'] = [f"{df['date'][idx][6:10]}" + \
                      f"-{df['qtr'][idx]}"
                      for idx in df.index]
    return df
            
            
#######################  Core Functions  ##############################
            
def data_reader(key_wrds, stop_wrds, file_addr, 
                wksht_name, first_col, last_col, 
                empty_cols=[], date_key=None, 
                column_names=None, columns_mdy=None):
    """_summary_
    
    This function uses read_wksht() and data_from_sheet()
    to return the block of data in a worksheet as a dataframe

    Args:
        key_wrd (_type_): _description_
        file_addr (_type_): _description_
        wksht_name (_type_): _description_
        empty_cols (_type_): _description_
        first_col (_type_): _description_
        last_col (_type_): _description_
        date_cell (_type_): _description_
        column_names (_type_): _description_

    Returns:
        _type_: _description_
    """
    
    wksht, name_date = read_wksht(file_addr, wksht_name, date_key)
    max_to_read = wksht.max_row
    
    # first data row to read. Follows key_wrd row.
    start = 1 + find_key_loc(1, max_to_read, wksht, key_wrds, first_col)
    # last data row to read. For rows, no key or row values.
    stop = -1 + find_key_loc(start, max_to_read, wksht, 
                             stop_wrds, first_col)
        
    data = data_from_sheet(wksht, first_col, last_col, empty_cols,
                           start, stop)
    data_df = pd.DataFrame(data, columns=column_names)
        
    # tidy 'date' col; create datetime and mdy cols
    # data_df['date'].apply(date_as_string) is too slow
    data_df['date'] = [date_as_string(item)
                       for item in data_df['date']]
    data_df['datetime'] = [datetime.strptime(item, '%m-%d-%Y').date()
                           for item in data_df['date']]
    
    # if any date is None, abort
    if any(item is None for item in list(data_df['date'])):
        print('In data_reader():')
        print('at least one date is None')
        print(key_wrds, file_addr, first_col, last_col)
        sys.exit()
        
    data_df[columns_mdy] = [item.split("-")
                            for item in data_df['date']]
    data_df[columns_mdy] = data_df[columns_mdy].astype('int16')
    
    # reduce (spurious) precision
    float64_cols = list(data_df.select_dtypes(include='float64').columns)
    data_df[float64_cols] = data_df[float64_cols].astype('float32')
    
    return [data_df, name_date]


def margin_reader(key_wrds, file_addr, wksht_name,
                  first_col):
    """read the operating margin data for S&P 500 by year.
       one df for each date of projections

    Returns:
        data_df   : pd.DateFrame containing margin data
        name_date : date of projection
    """
    
    wksht, _ = read_wksht(file_addr, wksht_name)
    max_to_read = wksht.max_row
    
    # find the rows with margin data
    start = 1 + find_key_loc(3, max_to_read, wksht, 
                             key_wrds, col=first_col)
    stop = start + 3
    
    # The label for the first col, which contains the qtr labels
    # is 'qtr'.  The other cols are named by year
    col_label = ["qtr"]
    max_to_read = 100
    l_col = -1 + find_key_loc(2, max_to_read, wksht,
                              [None], row=start-1)
    stop_col = col = openpyxl.utils.cell.get_column_letter(l_col)
    
    for step in range(2,l_col+1):
        col = openpyxl.utils.cell.get_column_letter(step)
        yr = str(wksht[f'{col}{start - 1}'].value)
        yr = "".join([ch for ch in yr if ch.isdigit()])
        col_label.append(yr)
    
    data = data_from_sheet(wksht, first_col, stop_col, [],
                           start, stop)
    data_df = pd.DataFrame(data, columns = col_label)
    
    # clean the entries in the 'qtr' col of the df
    data_df['qtr'] = [f'Q{4-q}' for q in range(0,4)]
    
    # reduce (spurious) precision
    float64_cols = list(data_df.select_dtypes(include='float64').columns)
    data_df[float64_cols] = data_df[float64_cols].astype('float32')

    return data_df


def  real_rt_reader(file_addr, name, prev_update):
    wkbk = load_workbook(filename=file_addr,
                         read_only=True,
                         data_only=True)
    wksht = wkbk.active
    min_row = 12
    max_row = wksht.max_row
    
    data = data_from_sheet(wksht, "A", "B", [],
                    min_row, max_row)
    
    df = pd.DataFrame(data, columns= ['datetime', name])
    
    # remove rows with N/A
    row_filter = df[name] != '#N/A'
    df = df[row_filter]
    df.reset_index(drop= True, inplace= True)
    
    # convert real rates to float32
    df[name] = df[name].astype('float32')
    
    # add year_qtr col, remove date-time
    date_lst = df['datetime'].dt.strftime('%Y-%m')
    df['year_qtr'] = \
        [f'{d[:4]}-Q{ ( (int(d[5:7]) - 1) // 3 ) + 1}'
         for d in date_lst]
    df = df[df.columns[1:]]

    return df


#=========================================================

if __name__ == '__main__':
    main()
# https://docs.python.org/3/library/__main__.html
# https://realpython.com/python-main-function/
    
